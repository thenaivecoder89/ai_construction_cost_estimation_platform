# File_name: boq_generation.py
# Purpose: Generate a detailed Bill of Quantities workbook from client RAG
# evidence, supported by corpus measurement and cost-estimating guidance.

import json
import re
import hashlib
from pathlib import Path
from datetime import datetime, timezone

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

from global_rag.scripts import config
import global_rag.scripts.retrieve_chunks as ret


BOQ_GENERATION_VERSION = "boq_generation_v1"
REPORT_TYPE = "ai_boq_generation"
CLASSIFICATION = "Confidential External"
WORKSTREAM = "ai_construction_cost_estimation_platform"

DIVISIONS = [
    ("03", "Concrete", "Div 03"),
    ("04", "Masonry", "Div 04"),
    ("05", "Metals", "Div 05"),
    ("06", "Wood, Plastic And Composites", "Div 06"),
    ("07", "Thermal And Moisture Protection", "Div 07"),
    ("08", "Openings", "Div 08"),
    ("09", "Finishes", "Div 09"),
    ("10", "Specialties", "Div 10"),
    ("12", "Furnishings", "Div 12"),
    ("14", "Conveying Equipment", "Div 14"),
    ("21", "Fire Suppression", "Div 21"),
    ("22", "Plumbing", "Div 22"),
    ("23", "Heating, Ventilating And Air Conditioning", "Div 23"),
    ("26", "Electrical", "Div 26"),
    ("27", "Communications", "Div 27"),
    ("28", "Electronic Safety And Security", "Div 28"),
]

DIVISION_BY_CODE = {code: {"name": name, "sheet": sheet} for code, name, sheet in DIVISIONS}

DEFAULT_UNITS = {
    "No": "Number",
    "No.": "Number",
    "nr": "Number",
    "lm": "Linear Meter",
    "Lm": "Linear Meter",
    "m": "Meter",
    "m2": "Square Meter",
    "m²": "Square Meter",
    "sqm": "Square Meter",
    "m3": "Cubic Meter",
    "m³": "Cubic Meter",
    "kg": "Kilogram",
    "ton": "Tonne",
    "L.S": "Lump Sum",
    "LS": "Lump Sum",
    "Item": "Composite Measurement",
}


def utc_now_iso():
    return datetime.now(timezone.utc).isoformat()


def clean_text(value):
    if value is None:
        return ""

    value = str(value).replace("\x00", " ")
    value = value.replace("\r", " ").replace("\n", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_text(value):
    return clean_text(value).lower()


def safe_float(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    value = clean_text(value)
    if value == "" or value in ["-", "Included", "Included if required", "N/A"]:
        return None

    try:
        if value.startswith("="):
            return None

        value = value.replace(",", "")
        value = value.replace("AED", "")
        value = value.strip()
        return float(value)
    except Exception:
        return None


def make_run_id(project_id):
    run_label = re.sub(r"[^A-Za-z0-9_]+", "_", clean_text(project_id)).strip("_")
    if not run_label:
        run_label = "PROJECT"

    raw_value = f"{run_label}_{utc_now_iso()}_{BOQ_GENERATION_VERSION}"
    digest = hashlib.sha256(raw_value.encode("utf-8")).hexdigest()[:12]
    return f"BOQ_{run_label}_{digest}"


def config_pack(project_id):
    config_base = config.config_base()
    config_paths = config.config_paths(client_data=project_id)
    boq_output_dir = Path(config_paths["output_dir"]) / "boq_generation"
    boq_output_dir.mkdir(parents=True, exist_ok=True)

    return {
        "config_base": config_base,
        "config_paths": config_paths,
        "boq_output_dir": boq_output_dir,
    }


def get_source_reference(result):
    source_reference = clean_text(result.get("source_reference"))
    return {
        "chunk_id": result.get("chunk_id"),
        "document_id": result.get("document_id"),
        "corpus_zone": result.get("corpus_zone"),
        "corpus_pack": result.get("corpus_pack"),
        "section_heading": result.get("section_heading"),
        "page_start": result.get("page_start"),
        "page_end": result.get("page_end"),
        "source_reference": source_reference,
    }


def run_retrieval(query_text, corpus_zone=None, corpus_pack=None, top_k=12):
    output = ret.retrieve_chunks(
        query_text=query_text,
        top_k=top_k,
        mode="hybrid",
        corpus_zone=corpus_zone,
        corpus_pack=corpus_pack,
        max_chunk_chars=8000,
    )
    return output.get("results", [])


def collect_boq_evidence(project_id):
    evidence_packets = {}

    retrieval_plan = {
        "client_boq": {
            "query": (
                "Building 1 BOQ bill of quantities item description quantity unit rate amount "
                "division concrete masonry finishes MEP electrical plumbing HVAC"
            ),
            "corpus_zone": "client_data",
            "corpus_pack": project_id,
            "top_k": 30,
        },
        "client_quantities": {
            "query": (
                "architectural drawings schedules wall partitions elevations plans sections quantities "
                "door schedule room finish schedule floor areas"
            ),
            "corpus_zone": "client_data",
            "corpus_pack": project_id,
            "top_k": 24,
        },
        "corpus_measurement_guidance": {
            "query": (
                "bill of quantities measurement rules NRM CESMM CSI divisions construction cost estimating "
                "quantity takeoff unit rate preliminaries contingency assumptions"
            ),
            "corpus_zone": "corpus_data",
            "corpus_pack": None,
            "top_k": 24,
        },
        "corpus_rate_benchmarks": {
            "query": (
                "UAE Dubai Abu Dhabi construction cost rates concrete masonry finishes MEP electrical "
                "plumbing HVAC cost guide price book benchmark"
            ),
            "corpus_zone": "corpus_data",
            "corpus_pack": None,
            "top_k": 24,
        },
    }

    for key, query_config in retrieval_plan.items():
        try:
            results = run_retrieval(
                query_text=query_config["query"],
                corpus_zone=query_config.get("corpus_zone"),
                corpus_pack=query_config.get("corpus_pack"),
                top_k=query_config.get("top_k", 12),
            )
            evidence_packets[key] = {
                "status": "ok",
                "query": query_config["query"],
                "results": results,
                "sources": [get_source_reference(item) for item in results],
                "evidence_blob": " ".join(clean_text(item.get("chunk_text")) for item in results),
            }
        except Exception as exc:
            evidence_packets[key] = {
                "status": "failed",
                "query": query_config["query"],
                "results": [],
                "sources": [],
                "evidence_blob": "",
                "error": f"{type(exc).__name__}: {str(exc)}",
            }

    return evidence_packets


def row_data_to_dict(row_data):
    if row_data is None:
        return {}

    if isinstance(row_data, dict):
        return row_data

    try:
        return json.loads(str(row_data))
    except Exception:
        return {}


def get_value_by_terms(row_dict, terms):
    for key, value in row_dict.items():
        key_text = normalize_text(key)
        if any(term in key_text for term in terms):
            return value

    return None


def ordered_row_values(row_dict):
    def sort_key(item):
        key = clean_text(item[0])
        match = re.search(r"(\d+)$", key)
        if match:
            return int(match.group(1))
        return 999

    return [value for _, value in sorted(row_dict.items(), key=sort_key)]


def infer_division_code(value, fallback_text=""):
    value = clean_text(value)
    fallback_text = clean_text(fallback_text)

    match = re.match(r"^(\d{2})$", value)
    if match and match.group(1) in DIVISION_BY_CODE:
        return match.group(1)

    match = re.search(r"\bdivision\s+(\d{2})\b", fallback_text, flags=re.IGNORECASE)
    if match and match.group(1) in DIVISION_BY_CODE:
        return match.group(1)

    match = re.search(r"\b(0[3-9]|1[024]|2[123678])\b", fallback_text)
    if match and match.group(1) in DIVISION_BY_CODE:
        return match.group(1)

    return None


def make_boq_item(
    division_code,
    section_code,
    item_code,
    description,
    unit,
    quantity=None,
    unit_rate=None,
    source="",
    confidence="medium",
):
    quantity_number = safe_float(quantity)
    unit_rate_number = safe_float(unit_rate)
    amount = None

    if quantity_number is not None and unit_rate_number is not None:
        amount = quantity_number * unit_rate_number

    return {
        "division_code": clean_text(division_code),
        "division_name": DIVISION_BY_CODE.get(clean_text(division_code), {}).get("name", ""),
        "section_code": clean_text(section_code),
        "item_code": clean_text(item_code),
        "description": clean_text(description),
        "unit": clean_text(unit),
        "quantity": quantity_number if quantity_number is not None else clean_text(quantity),
        "unit_rate_aed": unit_rate_number if unit_rate_number is not None else clean_text(unit_rate),
        "amount_aed": amount,
        "source": clean_text(source),
        "confidence": confidence,
    }


def fetch_client_boq_table_rows(project_id, db_url):
    engine = create_engine(db_url, pool_pre_ping=True)

    sql = text(
        """
        SELECT
            d.document_id,
            d.file_name,
            d.relative_path,
            t.table_id,
            t.table_name,
            t.sheet_name,
            tr.row_number,
            tr.row_data
        FROM extracted_table_rows tr
        JOIN extracted_tables t
            ON t.table_id = tr.table_id
        JOIN documents d
            ON d.document_id = tr.document_id
        WHERE d.corpus_zone = 'client_data'
          AND d.workstream = :workstream
          AND d.corpus_pack = :project_id
          AND (
                LOWER(d.file_name) LIKE '%boq%'
                OR LOWER(d.file_name) LIKE '%bill%quantit%'
                OR LOWER(t.table_name) LIKE '%boq%'
                OR LOWER(t.table_name) LIKE '%bill%quantit%'
              )
        ORDER BY
            d.document_id,
            t.table_id,
            tr.row_number;
        """
    )

    with engine.begin() as conn:
        rows = conn.execute(
            sql,
            {
                "workstream": WORKSTREAM,
                "project_id": project_id,
            },
        ).mappings().all()

    return [dict(row) for row in rows]


def extract_boq_items_from_table_rows(table_rows):
    items = []
    current_division_code = None
    current_section_code = None

    for row in table_rows:
        row_dict = row_data_to_dict(row.get("row_data"))
        values = ordered_row_values(row_dict)
        row_text = " ".join(clean_text(value) for value in values)

        if not row_text:
            continue

        inferred_division = infer_division_code(
            get_value_by_terms(row_dict, ["division", "div"]),
            fallback_text=f"{row.get('sheet_name')} {row.get('table_name')} {row_text}",
        )

        if inferred_division:
            current_division_code = inferred_division

        item_code = get_value_by_terms(row_dict, ["item", "activity id", "boq ref", "code"])
        description = get_value_by_terms(row_dict, ["description", "boq descri", "activity name"])
        unit = get_value_by_terms(row_dict, ["unit"])
        quantity = get_value_by_terms(row_dict, ["quantity", "qty"])
        rate = get_value_by_terms(row_dict, ["rate", "unit rate"])

        if len(values) >= 8:
            if current_division_code is None:
                current_division_code = infer_division_code(values[0], row_text)
            if item_code is None:
                item_code = values[2]
            if description is None:
                description = values[3]
            if unit is None:
                unit = values[4]
            if quantity is None:
                quantity = values[5]
            if rate is None:
                rate = values[6]

        item_code_text = clean_text(item_code)
        description_text = clean_text(description)
        unit_text = clean_text(unit)

        if current_division_code is None:
            continue

        if item_code_text == "" or description_text == "" or unit_text == "":
            if description_text and not item_code_text:
                section_match = re.search(r"\((\d{6})\)", description_text)
                if section_match:
                    current_section_code = section_match.group(1)
            continue

        if description_text.lower() in ["description", "boq descriotion", "boq description"]:
            continue

        section_code = current_section_code or clean_text(get_value_by_terms(row_dict, ["section"]))
        if not section_code:
            section_code = clean_text(values[1] if len(values) > 1 else "")

        items.append(
            make_boq_item(
                division_code=current_division_code,
                section_code=section_code,
                item_code=item_code_text,
                description=description_text,
                unit=unit_text,
                quantity=quantity,
                unit_rate=rate,
                source=(
                    f"document_id={row.get('document_id')}; table_id={row.get('table_id')}; "
                    f"sheet={row.get('sheet_name')}; row={row.get('row_number')}"
                ),
                confidence="high_client_boq_row",
            )
        )

    return items


def extract_boq_items_from_retrieved_chunks(evidence_packets):
    items = []
    text_blob = evidence_packets.get("client_boq", {}).get("evidence_blob", "")

    line_pattern = re.compile(
        r"\b(?P<div>0[3-9]|1[024]|2[123678])\s*[|,\- ]+"
        r"(?P<section>\d{2,6})?\s*[|,\- ]*"
        r"(?P<item>\d{3,5})\s*[|,\- ]+"
        r"(?P<description>.{8,180}?)\s*[|,\- ]+"
        r"(?P<unit>m2|m²|m3|m³|lm|Lm|No\.?|L\.S|LS|Item|kg|ton)\s*[|,\- ]+"
        r"(?P<qty>[0-9][0-9,\.]*)\s*[|,\- ]+"
        r"(?P<rate>[0-9][0-9,\.]*)",
        flags=re.IGNORECASE,
    )

    for match in line_pattern.finditer(text_blob):
        item = make_boq_item(
            division_code=match.group("div"),
            section_code=match.group("section") or "",
            item_code=match.group("item"),
            description=match.group("description"),
            unit=match.group("unit"),
            quantity=match.group("qty"),
            unit_rate=match.group("rate"),
            source="retrieved client BOQ chunk regex extraction",
            confidence="medium_retrieved_chunk",
        )
        items.append(item)

    return items


def build_minimum_scope_items():
    scope_seed = [
        ("03", "033000", "1000", "Cast-in-place reinforced concrete to structural elements", "m3"),
        ("04", "042000", "1000", "Concrete blockwork and masonry partitions", "m2"),
        ("05", "055000", "1000", "Metal fabrications and miscellaneous steelwork", "Item"),
        ("06", "062000", "1000", "Interior finish carpentry and joinery", "Item"),
        ("07", "071000", "1000", "Waterproofing and thermal/moisture protection", "m2"),
        ("08", "081000", "1000", "Doors, frames, glazing and opening assemblies", "No."),
        ("09", "092000", "1000", "Wall, ceiling and floor finishes", "m2"),
        ("10", "101000", "1000", "Specialties, signage and accessories", "Item"),
        ("12", "123000", "1000", "Counters, furnishings and built-in fittings", "Item"),
        ("14", "142000", "1000", "Elevators and conveying equipment", "L.S"),
        ("21", "211000", "1000", "Fire suppression systems", "Item"),
        ("22", "220000", "1000", "Plumbing systems", "Item"),
        ("23", "230000", "1000", "HVAC systems", "Item"),
        ("26", "260000", "1000", "Electrical systems", "Item"),
        ("27", "270000", "1000", "Communications systems", "Item"),
        ("28", "280000", "1000", "Electronic safety and security systems", "Item"),
    ]

    return [
        make_boq_item(
            division_code=division_code,
            section_code=section_code,
            item_code=item_code,
            description=description,
            unit=unit,
            quantity=None,
            unit_rate=None,
            source="minimum scope scaffold; quantities and rates require client evidence",
            confidence="low_scope_placeholder",
        )
        for division_code, section_code, item_code, description, unit in scope_seed
    ]


def compact_evidence_for_llm(evidence_packets, max_chars=30000):
    parts = []
    for packet_key, packet in evidence_packets.items():
        parts.append(f"\n[{packet_key}]")
        for result in packet.get("results", [])[:10]:
            source = get_source_reference(result)
            parts.append(
                f"Source {source.get('chunk_id')} {source.get('document_id')} "
                f"{source.get('corpus_zone')} {source.get('section_heading')}: "
                f"{clean_text(result.get('chunk_text'))[:1200]}"
            )

    return "\n".join(parts)[:max_chars]


def parse_json_from_response_text(value):
    value = clean_text(value)
    if not value:
        return None

    try:
        return json.loads(value)
    except Exception:
        pass

    match = re.search(r"(\{.*\})", value, flags=re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1))
        except Exception:
            return None

    return None


def generate_llm_boq_items(config_base, project_id, seed_items, evidence_packets, max_items=250):
    openai_api_key = config_base.get("openai_api_key")
    if not openai_api_key:
        return {
            "status": "skipped",
            "reason": "OPENAI_API_KEY is not configured.",
            "items": [],
        }

    client = OpenAI(api_key=openai_api_key)
    model = config_base.get("llm_model", "gpt-4.1-mini")

    seed_items_compact = [
        {
            "division_code": item.get("division_code"),
            "section_code": item.get("section_code"),
            "item_code": item.get("item_code"),
            "description": item.get("description"),
            "unit": item.get("unit"),
            "quantity": item.get("quantity"),
            "unit_rate_aed": item.get("unit_rate_aed"),
            "source": item.get("source"),
        }
        for item in seed_items[:max_items]
    ]

    prompt = f"""
You are a senior quantity surveyor generating a construction Bill of Quantities for {project_id}.

Use client evidence first. Use corpus evidence only to improve classification, measurement wording,
unit selection, benchmark flags, and completeness checks. Do not invent exact quantities or rates
where client evidence is missing.

Return JSON only with this shape:
{{
  "boq_items": [
    {{
      "division_code": "03",
      "section_code": "033000",
      "item_code": "1000",
      "description": "BOQ item description",
      "unit": "m2",
      "quantity": 123.0,
      "unit_rate_aed": 45.0,
      "source": "brief source note",
      "confidence": "high|medium|low"
    }}
  ],
  "quality_notes": ["note"]
}}

Seed BOQ rows:
{json.dumps(seed_items_compact, ensure_ascii=False)}

Retrieved evidence:
{compact_evidence_for_llm(evidence_packets)}
"""

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "system",
                "content": "Return only valid JSON. Do not use markdown fences.",
            },
            {
                "role": "user",
                "content": prompt,
            },
        ],
        max_output_tokens=16000,
        store=False,
    )

    response_text = getattr(response, "output_text", "")
    parsed = parse_json_from_response_text(response_text)
    if not parsed:
        return {
            "status": "failed",
            "reason": "LLM response did not contain valid JSON.",
            "items": [],
            "raw_response_preview": response_text[:1000],
        }

    normalized_items = []
    for item in parsed.get("boq_items", [])[:max_items]:
        division_code = clean_text(item.get("division_code"))
        if division_code not in DIVISION_BY_CODE:
            continue

        if not clean_text(item.get("description")) or not clean_text(item.get("unit")):
            continue

        normalized_items.append(
            make_boq_item(
                division_code=division_code,
                section_code=item.get("section_code"),
                item_code=item.get("item_code"),
                description=item.get("description"),
                unit=item.get("unit"),
                quantity=item.get("quantity"),
                unit_rate=item.get("unit_rate_aed"),
                source=item.get("source"),
                confidence=item.get("confidence", "medium_llm_synthesized"),
            )
        )

    return {
        "status": "ok",
        "items": normalized_items,
        "quality_notes": parsed.get("quality_notes", []),
    }


def dedupe_boq_items(items):
    deduped = []
    seen = set()

    for item in items:
        key = (
            clean_text(item.get("division_code")),
            clean_text(item.get("section_code")),
            clean_text(item.get("item_code")),
            normalize_text(item.get("description"))[:120],
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)

    return deduped


def group_items_by_division(items):
    grouped = {code: [] for code, _, _ in DIVISIONS}
    for item in items:
        division_code = clean_text(item.get("division_code"))
        if division_code in grouped:
            grouped[division_code].append(item)
    return grouped


def style_header_row(ws, row_number, start_col=1, end_col=8):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="B7B7B7")

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_money_cell(cell):
    cell.number_format = '#,##0.00'


def autosize_columns(ws, max_width=70):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[letter]:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


def create_summary_sheet(wb, grouped_items, run_id, project_id):
    ws = wb.create_sheet("Summary")
    ws["A1"] = "AI Generated Bill of Quantities"
    ws["A2"] = f"Project: {project_id}"
    ws["A3"] = f"Run ID: {run_id}"
    ws["A4"] = f"Generated at: {utc_now_iso()}"
    ws["A1"].font = Font(size=16, bold=True)

    start_row = 7
    headers = ["Division", "Description", "Line Items", "Total AED", "Notes"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=header)
    style_header_row(ws, start_row, 1, len(headers))

    row_number = start_row + 1
    for division_code, division_name, sheet_name in DIVISIONS:
        items = grouped_items.get(division_code, [])
        ws.cell(row=row_number, column=1, value=f"Division {division_code}")
        ws.cell(row=row_number, column=2, value=division_name)
        ws.cell(row=row_number, column=3, value=len(items))
        ws.cell(row=row_number, column=4, value=f"='{sheet_name}'!H{max(8, len(items) + 8)}")
        style_money_cell(ws.cell(row=row_number, column=4))
        ws.cell(row=row_number, column=5, value="Client evidence preferred; blanks require estimator review.")
        row_number += 1

    ws.cell(row=row_number + 1, column=3, value="Grand Total")
    ws.cell(row=row_number + 1, column=4, value=f"=SUM(D{start_row + 1}:D{row_number - 1})")
    ws.cell(row=row_number + 1, column=3).font = Font(bold=True)
    ws.cell(row=row_number + 1, column=4).font = Font(bold=True)
    style_money_cell(ws.cell(row=row_number + 1, column=4))
    autosize_columns(ws)


def create_abbreviations_sheet(wb):
    ws = wb.create_sheet("Abbreviations")
    ws["A1"] = "Units Of Measurement And Abbreviations"
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([])
    ws.append(["Abbreviation", "Meaning"])
    style_header_row(ws, 3, 1, 2)

    row_number = 4
    for unit, meaning in sorted(DEFAULT_UNITS.items()):
        ws.cell(row=row_number, column=1, value=unit)
        ws.cell(row=row_number, column=2, value=meaning)
        row_number += 1

    autosize_columns(ws)


def create_division_sheet(wb, division_code, division_name, sheet_name, items):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"DIVISION {division_code}"
    ws["A2"] = division_name.upper()
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"].font = Font(size=12, bold=True)

    headers = [
        "Div",
        "Section",
        "Item",
        "Description",
        "Unit",
        "Quantity",
        "Unit Rate AED",
        "Total Amount AED",
        "Source / Basis",
        "Confidence",
    ]

    header_row = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=header)
    style_header_row(ws, header_row, 1, len(headers))

    row_number = header_row + 1
    for item in items:
        ws.cell(row=row_number, column=1, value=item.get("division_code"))
        ws.cell(row=row_number, column=2, value=item.get("section_code"))
        ws.cell(row=row_number, column=3, value=item.get("item_code"))
        ws.cell(row=row_number, column=4, value=item.get("description"))
        ws.cell(row=row_number, column=5, value=item.get("unit"))
        ws.cell(row=row_number, column=6, value=item.get("quantity"))
        ws.cell(row=row_number, column=7, value=item.get("unit_rate_aed"))
        ws.cell(row=row_number, column=8, value=f"=F{row_number}*G{row_number}")
        ws.cell(row=row_number, column=9, value=item.get("source"))
        ws.cell(row=row_number, column=10, value=item.get("confidence"))

        style_money_cell(ws.cell(row=row_number, column=7))
        style_money_cell(ws.cell(row=row_number, column=8))
        ws.cell(row=row_number, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_number, column=9).alignment = Alignment(wrap_text=True, vertical="top")
        row_number += 1

    total_row = row_number + 1
    ws.cell(row=total_row, column=7, value="Division Total")
    ws.cell(row=total_row, column=8, value=f"=SUM(H{header_row + 1}:H{row_number - 1})")
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    style_money_cell(ws.cell(row=total_row, column=8))

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{header_row}:J{max(header_row, row_number - 1)}"
    autosize_columns(ws)
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["I"].width = 50


def create_cost_loading_sheet(wb, grouped_items):
    ws = wb.create_sheet("Cost Loading")
    headers = ["BOQ Ref", "Div", "Activity ID", "Activity Name", "BOQ Description", "QTY", "Unit", "Rate", "%", "Amount"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, 1, len(headers))

    row_number = 2
    grand_total_refs = []
    for division_code, division_name, _ in DIVISIONS:
        items = grouped_items.get(division_code, [])
        if not items:
            continue

        ws.cell(row=row_number, column=1, value=f"Division {division_code}")
        ws.cell(row=row_number, column=5, value=division_name)
        ws.cell(row=row_number, column=1).font = Font(bold=True)
        row_number += 1

        division_start = row_number
        for item in items:
            ws.cell(row=row_number, column=1, value=f"{division_code}-{item.get('item_code')}")
            ws.cell(row=row_number, column=2, value=division_code)
            ws.cell(row=row_number, column=3, value=item.get("item_code"))
            ws.cell(row=row_number, column=4, value=division_name)
            ws.cell(row=row_number, column=5, value=item.get("description"))
            ws.cell(row=row_number, column=6, value=item.get("quantity"))
            ws.cell(row=row_number, column=7, value=item.get("unit"))
            ws.cell(row=row_number, column=8, value=item.get("unit_rate_aed"))
            ws.cell(row=row_number, column=10, value=f"=F{row_number}*H{row_number}")
            style_money_cell(ws.cell(row=row_number, column=8))
            style_money_cell(ws.cell(row=row_number, column=10))
            row_number += 1

        ws.cell(row=row_number, column=9, value=f"Division {division_code} Total")
        ws.cell(row=row_number, column=10, value=f"=SUM(J{division_start}:J{row_number - 1})")
        style_money_cell(ws.cell(row=row_number, column=10))
        grand_total_refs.append(f"J{row_number}")
        row_number += 2

    ws.cell(row=row_number, column=9, value="Grand Total")
    ws.cell(row=row_number, column=10, value=f"=SUM({','.join(grand_total_refs)})" if grand_total_refs else 0)
    ws.cell(row=row_number, column=9).font = Font(bold=True)
    ws.cell(row=row_number, column=10).font = Font(bold=True)
    style_money_cell(ws.cell(row=row_number, column=10))
    ws.freeze_panes = "A2"
    autosize_columns(ws)
    ws.column_dimensions["E"].width = 60


def create_evidence_sheet(wb, evidence_packets, quality_notes):
    ws = wb.create_sheet("Evidence")
    ws["A1"] = "BOQ Evidence And Generation Notes"
    ws["A1"].font = Font(size=14, bold=True)

    ws.append([])
    ws.append(["Quality Notes"])
    style_header_row(ws, 3, 1, 1)
    row_number = 4
    for note in quality_notes:
        ws.cell(row=row_number, column=1, value=clean_text(note))
        row_number += 1

    row_number += 2
    headers = ["Evidence Packet", "Status", "Chunk ID", "Document ID", "Corpus Zone", "Corpus Pack", "Section", "Source Reference"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row_number, column=col, value=header)
    style_header_row(ws, row_number, 1, len(headers))
    row_number += 1

    for packet_key, packet in evidence_packets.items():
        if not packet.get("sources"):
            ws.cell(row=row_number, column=1, value=packet_key)
            ws.cell(row=row_number, column=2, value=packet.get("status"))
            ws.cell(row=row_number, column=8, value=packet.get("error", "No sources returned."))
            row_number += 1
            continue

        for source in packet.get("sources", [])[:20]:
            ws.cell(row=row_number, column=1, value=packet_key)
            ws.cell(row=row_number, column=2, value=packet.get("status"))
            ws.cell(row=row_number, column=3, value=source.get("chunk_id"))
            ws.cell(row=row_number, column=4, value=source.get("document_id"))
            ws.cell(row=row_number, column=5, value=source.get("corpus_zone"))
            ws.cell(row=row_number, column=6, value=source.get("corpus_pack"))
            ws.cell(row=row_number, column=7, value=source.get("section_heading"))
            ws.cell(row=row_number, column=8, value=source.get("source_reference"))
            row_number += 1

    autosize_columns(ws)
    ws.column_dimensions["H"].width = 70


def write_boq_workbook(output_path, run_id, project_id, items, evidence_packets, quality_notes):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    grouped_items = group_items_by_division(items)

    create_cost_loading_sheet(wb, grouped_items)
    create_summary_sheet(wb, grouped_items, run_id, project_id)
    create_abbreviations_sheet(wb)

    for division_code, division_name, sheet_name in DIVISIONS:
        create_division_sheet(
            wb=wb,
            division_code=division_code,
            division_name=division_name,
            sheet_name=sheet_name,
            items=grouped_items.get(division_code, []),
        )

    create_evidence_sheet(wb, evidence_packets, quality_notes)

    wb.save(output_path)


def summarize_items(items):
    grouped_items = group_items_by_division(items)
    division_summaries = []
    grand_total = 0.0

    for division_code, division_name, sheet_name in DIVISIONS:
        division_items = grouped_items.get(division_code, [])
        total = 0.0
        priced_items = 0

        for item in division_items:
            amount = safe_float(item.get("amount_aed"))
            if amount is not None:
                total += amount
                priced_items += 1

        grand_total += total
        division_summaries.append(
            {
                "division_code": division_code,
                "division_name": division_name,
                "sheet_name": sheet_name,
                "line_items": len(division_items),
                "priced_items": priced_items,
                "total_aed": total,
            }
        )

    return {
        "line_items": len(items),
        "priced_items": sum(row["priced_items"] for row in division_summaries),
        "grand_total_aed": grand_total,
        "division_summaries": division_summaries,
    }


def generate_boq(
    project_id,
    use_llm=False,
    write_workbook=True,
    max_llm_items=250,
):
    pack = config_pack(project_id)
    config_base = pack["config_base"]
    run_id = make_run_id(project_id)

    evidence_packets = collect_boq_evidence(project_id)

    direct_table_status = {
        "status": "not_attempted",
        "rows_found": 0,
        "items_extracted": 0,
        "error": None,
    }

    seed_items = []
    try:
        table_rows = fetch_client_boq_table_rows(
            project_id=project_id,
            db_url=config_base["db_url"],
        )
        direct_table_status["status"] = "ok"
        direct_table_status["rows_found"] = len(table_rows)
        seed_items = extract_boq_items_from_table_rows(table_rows)
        direct_table_status["items_extracted"] = len(seed_items)
    except Exception as exc:
        direct_table_status["status"] = "failed"
        direct_table_status["error"] = f"{type(exc).__name__}: {str(exc)}"

    if not seed_items:
        seed_items = extract_boq_items_from_retrieved_chunks(evidence_packets)

    if not seed_items:
        seed_items = build_minimum_scope_items()

    llm_status = {"status": "not_requested", "items": [], "quality_notes": []}
    if use_llm:
        llm_status = generate_llm_boq_items(
            config_base=config_base,
            project_id=project_id,
            seed_items=seed_items,
            evidence_packets=evidence_packets,
            max_items=max_llm_items,
        )

    if use_llm and llm_status.get("items"):
        items = llm_status["items"]
    else:
        items = seed_items

    items = dedupe_boq_items(items)
    summary = summarize_items(items)

    quality_notes = [
        "Client BOQ/extracted table rows are preferred over inferred RAG text.",
        "Corpus evidence is used for classification, measurement discipline and benchmark context; it does not override client quantities without evidence.",
        "Blank quantities or rates indicate unavailable client evidence and require estimator review.",
    ]

    if direct_table_status["items_extracted"] == 0:
        quality_notes.append(
            "No direct BOQ table rows were extracted from the database; workbook was generated from retrieved chunks or minimum scope scaffold."
        )

    if use_llm:
        quality_notes.extend(clean_text(note) for note in llm_status.get("quality_notes", []) if clean_text(note))

    output_files = {}
    if write_workbook:
        xlsx_path = pack["boq_output_dir"] / f"{run_id}.xlsx"
        write_boq_workbook(
            output_path=xlsx_path,
            run_id=run_id,
            project_id=project_id,
            items=items,
            evidence_packets=evidence_packets,
            quality_notes=quality_notes,
        )
        output_files["xlsx_path"] = str(xlsx_path)

    json_path = pack["boq_output_dir"] / f"{run_id}.json"
    output_payload = {
        "run_id": run_id,
        "project_id": project_id,
        "report_type": REPORT_TYPE,
        "classification": CLASSIFICATION,
        "boq_generation_version": BOQ_GENERATION_VERSION,
        "generated_at": utc_now_iso(),
        "direct_table_status": direct_table_status,
        "llm_status": {
            key: value
            for key, value in llm_status.items()
            if key != "items"
        },
        "summary": summary,
        "quality_notes": quality_notes,
        "items": items,
        "evidence_sources": {
            key: packet.get("sources", [])
            for key, packet in evidence_packets.items()
        },
    }
    json_path.write_text(json.dumps(output_payload, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    output_files["json_path"] = str(json_path)

    return {
        "message": "BOQ generation completed.",
        "status": "ok" if items else "validation_failed",
        "run_id": run_id,
        "project_id": project_id,
        "output_files": output_files,
        "summary": summary,
        "direct_table_status": direct_table_status,
        "llm_status": {
            key: value
            for key, value in llm_status.items()
            if key != "items"
        },
        "quality_notes": quality_notes,
    }


if __name__ == "__main__":
    print(
        json.dumps(
            generate_boq(
                project_id="ai_construction_cost_estimation_platform",
                use_llm=False,
                write_workbook=True,
            ),
            indent=2,
            ensure_ascii=False,
        )
    )
